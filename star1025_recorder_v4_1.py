import os, sys, datetime
import selenium as sel
import geckodriver_autoinstaller
from selenium import webdriver
import openpyxl as xl
from openpyxl import Workbook, load_workbook
import time
from datetime import date, datetime

from selenium.webdriver import Firefox
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile
from selenium.webdriver.common.by import By

geckodriver_autoinstaller.install()

daysOfWeek = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
_date = date.today()
timeElapsed = 0.0

def setupBrowser(browser):
	playButton = browser.find_element(By.CSS_SELECTOR,'.css-5c1va5')
	playButton.click()

def extractSiteData(browser):
	artist_name = browser.find_element(By.CSS_SELECTOR,'.css-1r4xrbt')
	song_name = browser.find_element(By.CSS_SELECTOR,'.css-4dhu2z')

	return [artist_name,song_name]

def generateCurrentFilepath(timestamp, _date):
	pathString = '/home/sdl5384/Desktop/' + str(_date.month) + "-" + str(_date.day)+'.xlsx'
	return pathString

def getTimestamp():
	stamp = time.strftime('%a %H %M %S')
	adjustedStamp = stamp.replace(':', ' ')
	return adjustedStamp

def createNewWorkbook(filepath):
	filepath = generateCurrentFilepath(timestamp, _date)
	workbook = Workbook()
	log = workbook.active
	log.title = getTimestamp()

	log['A2'].value = 'Date'
	log['B2'].value = 'Song Title'
	log['C2'].value = 'Artist'
	log['D2'].value = 'Time Played'

	workbook.save(filepath)

def recordData(filepath, musicData):
	#if file does exist
	daysOfWeek = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']

	if os.path.exists(filepath):
		#test if the program should create a new workbook
		#test if current month and current date are different than in filepath name
		if (_date.month != filepath[18:20] and _date.day != filepath[20:22]):
			timestamp = getTimestamp()
			path = generateCurrentFilepath(timestamp,_date)
			workbook = xl.load_workbook(path)
			ws = workbook.active

			ws['A2'].value = 'Date'
			ws['B2'].value = 'Song Title'
			ws['C2'].value = 'Artist'
			ws['D2'].value = 'Time Played'

			workbook.save(filepath)
			newRow = ws.max_row+1

			ws['A'+str(newRow)] = str(_date.month) + "/" + str(_date.day) + "/" + str(_date.year)
			ws['B'+str(newRow)] = musicData[0].text.replace('• Star 102.5','')
			ws['C'+str(newRow)] = musicData[1].text.replace('• Star 102.5','')
			ws['D'+str(newRow)] = time.strftime('%a %H %M %S')
			workbook.save(filepath)

		#if current month is the same but current date is different than in filepath name
		elif (_date.month != filepath[18:20] and _date.day == filepath[20:22]):
			createNewWorkbook(filepath)
		#if both current month and current date are the same
		elif (_date.month == filepath[18:20] and _date.day == filepath[20:22]):
			workbook = load_workbook(filepath)
			
			stamp = getTimestamp()
			workbook = Workbook()
			log = workbook.active
			log.title = getTimestamp()

			log['A2'].value = 'Date'
			log['B2'].value = 'Song Title'
			log['C2'].value = 'Artist'
			log['D2'].value = 'Time Played'

			workbook.save(filepath)
			newRow = log.max_row+1

			log['A'+str(newRow)] = str(_date.month) + "/" + str(_date.day) + "/" + str(_date.year)
			log['B'+str(newRow)] = musicData[0].text.replace('• Star 102.5','')
			log['C'+str(newRow)] = musicData[1].text.replace('• Star 102.5','')
			log['D'+str(newRow)] = time.strftime('%a %H %M %S') 			
			workbook.save(filepath)

	#if the file does not exist
	if not os.path.exists(filepath):
		timestamp = getTimestamp()
		filepath = generateCurrentFilepath(timestamp, _date)
		workbook = Workbook()
		log = workbook.active
		log.title = timestamp

		log['A2'].value = 'Date'
		log['B2'].value = 'Song Title'
		log['C2'].value = 'Artist'
		log['D2'].value = 'Time Played'

		workbook.save(filepath)
		newRow = log.max_row+1

		log['A'+str(newRow)] = str(_date.month) + "/" + str(_date.day) + "/" + str(_date.year)
		log['B'+str(newRow)] = musicData[0].text.replace('• Star 102.5','')
		log['C'+str(newRow)] = musicData[1].text.replace('• Star 102.5','')
		log['D'+str(newRow)] = time.strftime('%a %H %M %S') 
			
		workbook.save(filepath)

def determineApproximateSongLength(secondTimeInMin,secondTimeInSec,firstTimeInMin,firstTimeInSec):
	# diffInMin = secondTimeInMin - firstTimeInMin
	# diffInSec = secondTimeInSec - firstTimeInSec
	# minutesToSeconds = 
	return None

def main():
	m = _date.month
	d = _date.day

	url = 'https://www.audacy.com/stations/mystar1025'
	#profile = webdriver.FirefoxProfile()
	#profile.set_preference("network.http.phishy-userpass-length", 255)
	
	options = Options()
	options.set_preference('profile', url)
	service = Service('/usr/bin/geckodriver.exe')

	music_browser = Firefox(options=options)
	music_browser.get(url)

	time.sleep(10)
	setupBrowser(music_browser)
	time.sleep(10)

	currentData = extractSiteData(music_browser)
	if currentData[0].text == "Advertisement": 
		time.sleep(31)

	previousData = [None,None]

	i = 0
	secondsInMinute = 60
	minutesInHr = secondsInMinute
	secondsInHr = secondsInMinute*secondsInMinute
	timeCheckInterval = 3
	runningTime = 7
	
	#run while loop for 7 hours of time (during the night to collect overight data)
	#code will test every three seconds in change of currently playing song
	while i <= ((runningTime*secondsInHr)/timeCheckInterval):
		if previousData[0] != currentData[0].text and previousData[1] != currentData[1].text:
			temp1 = currentData[0].text
			temp2 = currentData[1].text
			previousData[0] = temp1
			previousData[1] = temp2

			recordData('/home/sdl5384/Desktop/'+str(m)+"-"+str(d)+'.xlsx', currentData)
			print(f"New Song: {currentData[0].text}, {currentData[1].text}")
			print(f"Current time run: {getTimestamp()}")
		if previousData[0] == currentData[0].text and previousData[1] == currentData[1].text:
			time.sleep(timeCheckInterval)
	
		i = i + 1

	music_browser.quit()

main()