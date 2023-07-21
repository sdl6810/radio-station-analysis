import glob
from openpyxl import load_workbook, Workbook
import statistics

def getStarDataFiles(filepathRoot):	
	firstList = glob.glob(filepathRoot + '/[0-9][0-9]' + '-' + '[0-9][0-9].xlsx')
	secondList = glob.glob(filepathRoot + '/[0-9][0-9]' + '-' + '[0-9].xlsx')
	finalList = []

	for i in range(0, len(firstList)):
		finalList.append(firstList[i])

	for j in range(0, len(secondList)):
		finalList.append(secondList)

	return finalList

def aggregateData(files):
	workbook = Workbook()
	log = workbook.active
	log.title = "Data"

	log.cell(2,2).value = 'Date'
	log.cell(2,3).value = 'Deejay'
	log.cell(2,4).value = 'Song Title'
	log.cell(2,5).value = 'Artist'
	log.cell(2,6).value = 'Time'
	#for i in range(0, len(files)):

	workbook.save('/home/sdl5384/Desktop/Music/dataDump22.xlsx')

def main():
	path = '/home/sdl5384/Desktop/data2022'

	f = getStarDataFiles(path)
	aggregateData(f)

main()