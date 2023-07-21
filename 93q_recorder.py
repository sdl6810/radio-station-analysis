import os, sys, datetime
import selenium as sel
import geckodriver_autoinstaller
from selenium import webdriver
import openpyxl as xl
from openpyxl import Workbook, load_workbook
import time
from datetime import date, datetime

from selenium.webdriver import Firefox
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile
from selenium.webdriver.common.by import By

geckodriver_autoinstaller.install()

daysOfWeek = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
timestamp = time.strftime('%a %H:%M:%S')
_date = date.today()

def extractSiteData(browser):
	raw = browser.find_element(By.CSS_SELECTOR,'.tablelist-schedule').text
	raw_text = raw.split('\n')

	return raw_text

def generateCurrentFilepath(timestamp, _date):
	pathString = '/home/sdl5384/Desktop/93q ' + str(_date.month) + "-" + str(_date.day)+'.xlsx'
	return pathString

def getTimestamp():
	stamp = time.strftime('%a %H %M %S')
	adjustedStamp = stamp.replace(':', ' ')

	return adjustedStamp

def organizeMusicData(musicData, i):
	timePlayed = musicData[i][0:5]
	songAndArtist = musicData[i][5:len(musicData[i])]
	songArtist_list = songAndArtist.split(' - ')

	return [timePlayed, songArtist_list[0], songArtist_list[1]]


def createNewWorkbook(filepath):
	filepath = generateCurrentFilepath(timestamp, _date)
	workbook = Workbook()
	log = workbook.active
	log.title = getTimestamp()

	log['A2'].value = 'Date'
	log['B2'].value = 'DJ on shift'
	log['C2'].value = 'Song Title'
	log['D2'].value = 'Artist'
	log['E2'].value = 'Time Played'

	workbook.save(filepath)
	return workbook

def recordData(filepath, musicData):
	#if file does exist
	daysOfWeek = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']

	#dj = determineDJ()

	if os.path.exists(filepath):
		#test if the program should create a new workbook
		#test if current month and current date are different than in filepath name
		if (_date.month != filepath[18:20] and _date.day != filepath[20:22]):
			path = generateCurrentFilepath(timestamp,_date)
			workbook = xl.load_workbook(filepath)

			newSheet = workbook.create_sheet(getTimestamp())

			newSheet['A2'].value = 'Date'
			newSheet['B2'].value = 'DJ on shift'
			newSheet['C2'].value = 'Song Title'
			newSheet['D2'].value = 'Artist'
			newSheet['E2'].value = 'Time Played'
			workbook.save(filepath)

			i = 0
			while (i <= len(musicData)-1):
				data = organizeMusicData(musicData,i)
				#item 0 is time played
				#item 1 is artist
				#item 2 is song

				newSheet['A'+str(3+i)].value = str(_date.month) + "/" + str(_date.day) + "/" + str(_date.year)
				newSheet['B'+str(3+i)].value = ''
				newSheet['C'+str(3+i)].value = data[2]
				newSheet['D'+str(3+i)].value = data[1]
				newSheet['E'+str(3+i)].value = data[0]
				workbook.save(filepath)
				i=i+1
				
			workbook.save(filepath)

		#if current month is the same but current date is different than in filepath name
		elif (_date.month != filepath[18:20] and _date.day == filepath[20:22]):
			createNewWorkbook(filepath)
		#if both current month and current date are the same
		elif (_date.month == filepath[18:20] and _date.day == filepath[20:22]):
			workbook = load_workbook(filepath)
			
			stamp = getTimestamp()
			workbook = Workbook()
			log = workbook.active
			log.title = getTimestamp()

			log['A2'].value = 'Date'
			log['B2'].value = 'DJ on shift'
			log['C2'].value = 'Song Title'
			log['D2'].value = 'Artist'
			log['E2'].value = 'Time Played'

			data = organizeMusicData(musicData)
			#item 0 is time played
			#item 1 is artist
			#item 2 is song

			i = 0
			while (i <= len(musicData)-1):
				data = organizeMusicData(musicData,i)
				#item 0 is time played
				#item 1 is artist
				#item 2 is song

				newSheet['A'+str(3+i)].value = str(_date.month) + "/" + str(_date.day) + "/" + str(_date.year)
				newSheet['B'+str(3+i)].value = ''
				newSheet['C'+str(3+i)].value = data[2]
				newSheet['D'+str(3+i)].value = data[1]
				newSheet['E'+str(3+i)].value = data[0]
				i=i+1
			
			workbook.save(filepath)

	#if the file does not exist
	if not os.path.exists(filepath):
		workbook = createNewWorkbook(filepath)
		log = workbook.active
		#log.title = getTimestamp()

		log['A2'].value = 'Date'
		log['B2'].value = 'DJ on shift'
		log['C2'].value = 'Song Title'
		log['D2'].value = 'Artist'
		log['E2'].value = 'Time Played'
		workbook.save(filepath)

		i = 0
		while (i <= len(musicData)-1):
			data = organizeMusicData(musicData,i)
			#item 0 is time played
			#item 1 is artist
			#item 2 is song

			log['A'+str(3+i)].value = str(_date.month) + "/" + str(_date.day) + "/" + str(_date.year)
			log['B'+str(3+i)].value = ''
			log['C'+str(3+i)].value = data[2]
			log['D'+str(3+i)].value = data[1]
			log['E'+str(3+i)].value = data[0]
			i=i+1
			
			workbook.save(filepath)


def emailSpreadsheet(filepath, email):
	return None

def main():
	m = _date.month
	d = _date.day

	url = 'https://onlineradiobox.com/us/wntq/playlist/?cs=us.wntq'
	#profile = webdriver.FirefoxProfile()
	#profile.set_preference("network.http.phishy-userpass-length", 255)
	
	options = Options()
	options.set_preference('profile', url)
	service = Service('/usr/bin/geckodriver.exe')

	music_browser = Firefox(options=options)
	music_browser.get(url)

	data = extractSiteData(music_browser)

	recordData('/home/sdl5384/Desktop/93q '+str(m)+"-"+str(d)+'.xlsx', data)
	music_browser.quit()

main()