import os, sys, datetime
import selenium as sel
import geckodriver_autoinstaller
from selenium import webdriver
import openpyxl as xl
from openpyxl import Workbook, load_workbook
import time
from datetime import date, datetime

from selenium.webdriver import Firefox
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile
from selenium.webdriver.common.by import By

geckodriver_autoinstaller.install()

daysOfWeek = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
timestamp = time.strftime('%a %H:%M:%S')
_date = date.today()

def setupBrowser(browser):	
	playButton = browser.find_element_by_id("playButton")
	muteButton = browser.find_element_by_id("muteButton")

def extractSiteData(browser):
	raw = browser.find_element(By.CSS_SELECTOR,'.station-recently-played')
	raw_list = raw.text.split('\n')

	return raw_list

def generateCurrentFilepath(timestamp, _date):
	pathString = '/home/sdl5384/Desktop/' + str(_date.month) + "-" + str(_date.day)+' b101philly.xlsx'
	return pathString

def getTimestamp():
	stamp = time.strftime('%a %H %M %S')
	adjustedStamp = stamp.replace(':', ' ')
	return adjustedStamp

def createNewWorkbook(filepath):
	filepath = generateCurrentFilepath(timestamp, _date)
	workbook = Workbook()
	log = workbook.active
	log.title = getTimestamp()

	log['A2'].value = 'Date'
	#log['B2'].value = "Deejay"
	log['C2'].value = 'Song Title'
	log['D2'].value = 'Artist'
	log['E2'].value = 'Time Played'

	workbook.save(filepath)

def recordData(filepath, musicData):
	#if file does exist
	daysOfWeek = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']

	#dj = determineDJ()

	if os.path.exists(filepath):
		#test if the program should create a new workbook
		#test if current month and current date are different than in filepath name
		if (_date.month != filepath[18:20] and _date.day != filepath[20:22]):
			path = generateCurrentFilepath(timestamp,_date)
			workbook = xl.load_workbook(path)

			newSheet = workbook.create_sheet(getTimestamp())

			newSheet['A2'].value = 'Date'
			#newSheet['B2'].value = "Deejay"
			newSheet['C2'].value = 'Song Title'
			newSheet['D2'].value = 'Artist'
			newSheet['E2'].value = 'Time Played'

			i = 0
			while (i <= (len(musicData)/3)-1):
				newSheet['A'+str(3+i)].value = str(_date.month) + "/" + str(_date.day) + "/" + str(_date.year)
				#newSheet['B'+str(3+i)].value = dj
				newSheet['E'+str(3+i)].value = musicData[3*i]	#time played: 0, 3, 6, 9, 12, ..., 3i
				newSheet['C'+str(3+i)].value = musicData[(3*i)+1]	#song played: 1, 4, 7, 10, 13, ...3i+1
				newSheet['D'+str(3+i)].value = musicData[(3*i)+2]	#artist played: 2, 5, 8, 11, 14, ...,3i+2
				i=i+1
				
				workbook.save(filepath)

		#if current month is the same but current date is different than in filepath name
		elif (_date.month != filepath[18:20] and _date.day == filepath[20:22]):
			createNewWorkbook(filepath)
		#if both current month and current date are the same
		elif (_date.month == filepath[18:20] and _date.day == filepath[20:22]):
			workbook = load_workbook(filepath)
			
			stamp = getTimestamp()
			workbook = Workbook()
			log = workbook.active
			log.title = getTimestamp()

			log['A2'].value = 'Date'
			#log['B2'].value = "Deejay"
			log['C2'].value = 'Song Title'
			log['D2'].value = 'Artist'
			log['E2'].value = 'Time Played'

			i = 0
			while (i <= (len(musicData)/3)-1):
				newSheet['A'+str(3+i)].value = str(_date.month) + "/" + str(_date.day) + str(_date.year)
				#newSheet['B'+str(3+i)].value = dj
				newSheet['E'+str(3+i)].value = musicData[3*i]	#time played: 0, 3, 6, 9, 12, ..., 3i
				newSheet['C'+str(3+i)].value = musicData[(3*i)+1]	#song played: 1, 4, 7, 10, 13, ...3i+1
				newSheet['D'+str(3+i)].value = musicData[(3*i)+2]	#artist played: 2, 5, 8, 11, 14, ...,3i+2
				i=i+1
			
				workbook.save(filepath)

	#if the file does not exist
	if not os.path.exists(filepath):
		filepath = generateCurrentFilepath(timestamp, _date)
		stamp = getTimestamp()
		workbook = Workbook()
		log = workbook.active
		log.title = getTimestamp()

		log['A2'].value = 'Date'
		#log['B2'].value = 'Deejay'
		log['C2'].value = 'Song Title'
		log['D2'].value = 'Artist'
		log['E2'].value = 'Time Played'

		i = 0
		while (i <= (len(musicData)/3)-1):
			log['A'+str(3+i)].value = str(_date.month) + "/" + str(_date.day) + "/" + str(_date.year)
			#log['B'+str(3+i)].value = dj
			log['E'+str(3+i)].value = musicData[3*i]	#time played: 0, 3, 6, 9, 12, ..., 3i
			log['C'+str(3+i)].value = musicData[(3*i)+1]	#song played: 1, 4, 7, 10, 13, ...3i+1
			log['D'+str(3+i)].value = musicData[(3*i)+2]	#artist played: 2, 5, 8, 11, 14, ...,3i+2
			i=i+1
			
			workbook.save(filepath)

# def weekdayDeejay(hour):
# 	deejay = ''
# 	if (0 <= hour <= 6):
# 		deejay = 'STAR music'
# 	elif (6 <= hour <=12):
# 		deejay = 'Rob Lucas'
# 	elif (12 <= hour <= 18):
# 		deejay = 'Sue ONeil'
# 	else:
# 		deejay = 'STAR Music'

# 	return deejay

# def saturdayDeejay(hour):
# 	deejay = ''
# 	if (9 <= hour <= 10):
# 		deejay = 'Mike McQueen'
# 	elif (10 <= hour <= 14):
# 		deejay = 'Rob Lucas'
# 	elif (14 <= hour <= 18):
# 		deejay = 'Sue ONeil'
# 	else:
# 		deejay = 'STAR Music'

# 	return deejay

# def sundayDeejay():
# 	return 'STAR Music'

# def determineDJ():
# 	deejay = ''
# 	daysOfWeek = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']

# 	stamp = time.strftime("%a %H:%M:%S")
# 	wkday = stamp[0:3]
# 	wkday_time = stamp.replace(stamp[0:4],'')
# 	hrStamp = datetime.today().hour

# 	dayNumber = daysOfWeek.index(wkday)
# 	if 1 <= dayNumber <= 5:
# 		deejay = weekdayDeejay(hrStamp)
# 		print(deejay)
# 	elif dayNumber == 0:
# 		deejay = sundayDeejay()
# 		print(deejay)
# 	elif dayNumber == 6:
# 		deejay = saturdayDeejay(hrStamp)
# 		print(deejay)

# 	return deejay

def emailSpreadsheet(filepath, email):
	return None

def main():
	m = _date.month
	d = _date.day

	url = 'https://www.audacy.com/b101philly/listen#recently-played'
	#profile = webdriver.FirefoxProfile()
	#profile.set_preference("network.http.phishy-userpass-length", 255)
	
	options = Options()
	options.set_preference('profile', url)
	service = Service('/usr/bin/geckodriver.exe')

	music_browser = Firefox(options=options)
	music_browser.get(url)

	music_browser.refresh()

	music_browser.refresh()

	data = extractSiteData(music_browser)

	recordData('/home/sdl5384/Desktop/'+str(m)+"-"+str(d)+' b101philly.xlsx', data)
	music_browser.quit()

main()
