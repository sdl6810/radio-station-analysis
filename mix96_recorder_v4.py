#!/usr/bin/python3

import os, sys, datetime
import selenium as sel
from selenium import webdriver
import openpyxl as xl
from openpyxl import Workbook, load_workbook
import time
from datetime import date, datetime
import geckodriver_autoinstaller

from selenium.webdriver import Firefox
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile
from selenium.webdriver.common.by import By

daysOfWeek = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
timestamp = time.strftime('%a %H:%M:%S')
_date = date.today()

def setupBrowser(browser):	
	playButton = browser.find_element(By.CSS_SELECTOR,'#playButton')
	playButton.click()

def extractSiteData(browser):
	raw_list = []
	print(raw_list)
	i = 1
	while (i <= 21):
		artist_title = browser.find_element(By.CSS_SELECTOR,'section.recently-played-list > article:nth-child('+str(i)+') > a:nth-child(2) > div:nth-child(1) > div:nth-child(1)').text
		time_played = browser.find_element(By.CSS_SELECTOR,'section.recently-played-list > article:nth-child('+str(i)+') > a:nth-child(2) > div:nth-child(1) > div:nth-child(2)').text
		if artist_title[len(artist_title)-1] == "-":
			raw_list.append(artist_title)
			raw_list.append('NO SONG TITLE LISTED')
			raw_list.append(time_played)
		else:
			raw_list.append(artist_title.split(' - ')[0])
			raw_list.append(artist_title.split(' - ')[1])
			raw_list.append(time_played)
		i = i+1

	return raw_list

def generateCurrentFilepath(timestamp,_date):
	pathString = '/home/sdl5384/Desktop/mix96_' + str(_date.month) + "-" + str(_date.day)+'.xlsx'
	return pathString

def getTimestamp():
	stamp = time.strftime('%a %H %M %S')
	adjustedStamp = stamp.replace(':', ' ')
	return adjustedStamp

def createNewWorkbook(filepath):
	filepath = generateCurrentFilepath(timestamp, _date)
	workbook = Workbook()
	log = workbook.active
	log.title = getTimestamp()

	log['A2'].value = 'Date'
	log['B2'].value = 'DJ on shift'
	log['C2'].value = 'Song Title'
	log['E2'].value = 'Artist'
	log['D2'].value = 'Time Played'

	workbook.save(filepath)

def recordData(filepath, musicData):
	#if file does exist
	daysOfWeek = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']

	if os.path.exists(filepath):
		#test if the program should create a new workbook
		#test if current month and current date are different than in filepath name
		if (_date.month != filepath[18:20] and _date.day != filepath[20:22]):
			path = generateCurrentFilepath(timestamp,_date)
			workbook = xl.load_workbook(path)

			newSheet = workbook.create_sheet(getTimestamp())

			newSheet['A2'].value = 'Date'
			newSheet['B2'].value = 'Hour'
			newSheet['C2'].value = 'Song Title'
			newSheet['E2'].value = 'Artist'
			newSheet['D2'].value = 'Time Played'

			i = 0
			while (i <= (len(musicData)/3)-1):
				newSheet['A'+str(3+i)].value = str(_date.month) + "/" + str(_date.day) + "/" + str(_date.year)
				newSheet['B'+str(3+i)].value = ''
				newSheet['E'+str(3+i)].value = musicData[3*i]	#time played: 0, 3, 6, 9, 12, ..., 3i
				newSheet['C'+str(3+i)].value = musicData[(3*i)+1]	#song played: 1, 4, 7, 10, 13, ...3i+1
				newSheet['D'+str(3+i)].value = musicData[(3*i)+2]	#artist played: 2, 5, 8, 11, 14, ...,3i+2
				i=i+1
				
				workbook.save(filepath)

		#if current month is the same but current date is different than in filepath name
		elif (_date.month != filepath[18:20] and _date.day == filepath[20:22]):
			createNewWorkbook(filepath)
		#if both current month and current date are the same
		elif (_date.month == filepath[18:20] and _date.day == filepath[20:22]):			
			workbook = Workbook()
			log = workbook.active
			log.title = getTimestamp()

			log['A2'].value = 'Date'
			log['B2'].value = 'Hour'
			log['C2'].value = 'Song Title'
			log['E2'].value = 'Artist'
			log['D2'].value = 'Time Played'

			i = 0
			while (i <= (len(musicData)/3)-1):
				newSheet['A'+str(3+i)].value = str(_date.month) + "/" + str(_date.day) + "/" + str(_date.year)
				newSheet['B'+str(3+i)].value = ''
				newSheet['E'+str(3+i)].value = musicData[3*i]	#time played: 0, 3, 6, 9, 12, ..., 3i
				newSheet['C'+str(3+i)].value = musicData[(3*i)+1]	#song played: 1, 4, 7, 10, 13, ...3i+1
				newSheet['D'+str(3+i)].value = musicData[(3*i)+2]	#artist played: 2, 5, 8, 11, 14, ...,3i+2
				i=i+1
			
				workbook.save(filepath)

	#if the file does not exist
	if not os.path.exists(filepath):
		filepath = generateCurrentFilepath(timestamp, _date)
		workbook = Workbook()
		log = workbook.active
		log.title = getTimestamp()

		log['A2'].value = 'Date'
		log['B2'].value = 'Hour'
		log['C2'].value = 'Song Title'
		log['E2'].value = 'Artist'
		log['D2'].value = 'Time Played'

		i = 0
		while (i <= (len(musicData)/3)-1):
			log['A'+str(3+i)].value = str(_date.month) + "/" + str(_date.day) + "/" + str(_date.year)
			log['B'+str(3+i)].value = ''
			log['E'+str(3+i)].value = musicData[3*i]	#time played: 0, 3, 6, 9, 12, ..., 3i
			log['C'+str(3+i)].value = musicData[(3*i)+1]	#song played: 1, 4, 7, 10, 13, ...3i+1
			log['D'+str(3+i)].value = musicData[(3*i)+2]	#artist played: 2, 5, 8, 11, 14, ...,3i+2
			i=i+1
			
			workbook.save(filepath)

def main():
	m = _date.month
	d = _date.day

	url = 'https://961thebreeze.com/playlist/'
	geckodriver_autoinstaller.install()
	
	options = Options()
	options.set_preference('profile', url)
	service = Service('/usr/bin/geckodriver.exe')

	music_browser = Firefox(options=options)
	music_browser.get(url)

	data = extractSiteData(music_browser)

	recordData('/home/sdl5384/Desktop/mix96_'+str(m)+"-"+str(d)+'.xlsx', data)
	music_browser.quit()

main()
